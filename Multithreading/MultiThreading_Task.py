#importing libraries.
import openpyxl

#importing the dataset.
path="multi_threading_activity.xlsx"

obj = openpyxl.load_workbook(path)

sheet1 = obj.active

mat=[]

for i in range(1, sheet1.max_row+1):
  row=[]
  for j in range(1, sheet1.max_column+1):
    cell_obj = sheet1.cell(row=i, column=j)
    row.append(cell_obj.value)
  mat.append(row)
 
mat.pop()

mat.pop()

new=mat[1:]

for i in range(len(new)):
  for j in range(2):
    if  type(new[i][j])==str:
      print(new[i][j])
      new[i][j]=0
    if  type(new[i][j])==None:
      print(new[i][j])
      new[i][j]=0

print(new)      

print(len(new))

import threading

result=[]
def sum_of_two_rows(new):
  sumMat=[]
  for i in range(len(new)):
    sumMat.append(new[i][0]+new[i][1])
  result.append(sumMat) 
  print(sumMat)

def diff_of_tow_rows(new):
  diffMat=[]
  for i in range(len(new)):
    diffMat.append(new[i][0]-new[i][1])
  result.append(diffMat)
  print(diffMat)

t1=threading.Thread(target=sum_of_two_rows,args=(new,))

t2=threading.Thread(target=diff_of_tow_rows,args=(new,))

t1.start()

t2.start()

a=t1.join()

b=t2.join()

print(result)

col1=result[0]

col1

col2=result[1]

