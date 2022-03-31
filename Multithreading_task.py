#!/usr/bin/env python
# coding: utf-8

# In[2]:


import openpyxl
obj = openpyxl.load_workbook("multi_threading_activity.xlsx")


# In[3]:


task1 = obj.active


# In[4]:


mat=[]


# In[5]:


for i in range(1, task1.max_row+1):
    row=[]
    for j in range(1, task1.max_column+1):
        cell_obj = task1.cell(row=i, column=j)
        row.append(cell_obj.value)
        mat.append(row)
print(mat)  


# In[ ]:


mat.pop()


# In[ ]:


mat.pop()


# In[ ]:


new=mat[1:]


# In[ ]:


for i in range(len(new)):
      for j in range(2):
            if  type(new[i][j])==str:
                 print(new[i][j])
                    new[i][j]=0
                    
            if  type(new[i][j])==None:
                 print(new[i][j])
                    new[i][j]=0
                
            
print(new)      


# In[ ]:


print(len(new))


# In[ ]:


import threading 


# In[ ]:


def sum_of_two_rows(new):
    sumMat=[]
    for i in range(len(new)):

    if new[i][0] and new[i][1]==None:
    sumMat.append(0)
    
    sumMat.append(new[i][0]+new[i][1])

print(sumMat)


# In[ ]:


def diff_of_tow_rows(new):
    diffMat=[]
    for i in range(len(new)):
    if new[i][0] and new[i][1]==None:
    diffMat.append(0)
    diffMat.append(new[i][0]-new[i][1])

print(diffMat)


# In[ ]:


t1=threading.Thread(target=sum_of_two_rows,args=(new,))


# In[ ]:


t2=threading.Thread(target=diff_of_tow_rows,args=(new,))


# In[ ]:


t1.start()


# In[ ]:


t2.start()

